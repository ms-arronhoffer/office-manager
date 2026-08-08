"""Import router: template downloads and XLSX import uploads."""

from __future__ import annotations

import io
from fastapi import APIRouter, Depends, Query, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.database import get_db
from app.auth.dependencies import require_role
from app.models.import_batch import ImportBatch, content_fingerprint
from app.models.user import User
from app.services import import_assurance
from app.services.import_service import IMPORTERS

router = APIRouter()

# ── Template definitions ──────────────────────────────────────────────

TEMPLATES: dict[str, dict] = {
    "managers": {
        "filename": "managers_template.xlsx",
        "title": "Managers",
        "headers": ["Name", "Email", "Phone"],
        "example": ["John Smith", "jsmith@example.com", "555-0100"],
    },
    "offices": {
        "filename": "offices_template.xlsx",
        "title": "Offices",
        "headers": [
            "Office Number", "Region Number", "Location Type", "Location Name",
            "Manager", "Active", "Address Line 1", "Address Line 2", "City",
            "State", "Zip Code", "Phone", "Fax", "Email", "Mail/Shipping",
            "Sector", "Notes", "Other Names", "Crown Property On Site",
            "Additional Info", "Closing Notes", "Total SqFt", "Usable SqFt",
            "Headcount Capacity", "Current Headcount", "Space Type",
            "GL Account Code",
            "Owner Same As Landlord", "Owner Name", "Owner Company",
            "Owner Email", "Owner Phone", "Owner Address Line 1",
            "Owner Address Line 2", "Owner City", "Owner State", "Owner Zip Code",
        ],
        "example": [
            101, 1, "office", "Main Office", "John Smith", "Yes",
            "123 Main St", "Suite 200", "Springfield", "IL", "62701",
            "555-0100", "555-0101", "main@example.com", "Same as office",
            "Government", "Example office", "Downtown Branch", "None",
            "Renovated 2022", "", 12000, 9500, 60, 45, "office",
            "6000",
            "No", "Property Owner LLC", "Owner Holdings Inc",
            "owner@example.com", "555-0102", "1 Owner Way", "Floor 3",
            "Chicago", "IL", "60601",
        ],
    },
    "leases": {
        "filename": "leases_template.xlsx",
        "title": "Leases",
        "headers": [
            "Lease Name", "Office Number", "Manager", "Expiration Date",
            "Lessor Name", "Notice Period", "Notice Days", "Notice Date",
            "Notice Given Date", "Status", "Expiration Year",
            "Lease Commencement Date", "Accounting Standard",
            "Lease Classification", "Payment Amount", "Payment Frequency",
            "Annual Escalation Rate", "Incremental Borrowing Rate",
            "Initial Direct Costs", "Lease Incentives", "Prepaid Rent",
            "Residual Value Guarantee", "Short Term Lease", "Low Value Lease",
            "Currency",
        ],
        "example": [
            "101 - Main Office", 101, "John Smith", "12/31/2026",
            "ABC Properties", "90 Days", 90, "10/02/2026",
            "", "Active", 2026,
            "01/01/2021", "asc842", "operating", 5000, "monthly",
            0.03, 0.05, 2500, 1000, 5000, 0, "No", "No", "USD",
        ],
    },
    "landlords": {
        "filename": "landlords_template.xlsx",
        "title": "Landlords",
        "headers": [
            "ERN", "Office Name", "Office Number", "Landlord Company",
            "Contact Name", "Title", "Email", "Phone", "Mailing Address",
            "Online Sign In", "Vendor ID", "Notes", "Address",
            "Address Line 1", "Address Line 2", "City", "State", "Zip Code",
            "Mailing Address Line 1", "Mailing Address Line 2", "Mailing City",
            "Mailing State", "Mailing Zip Code", "Secondary Phone", "Fax",
            "Website", "Entity Type", "Tax ID", "Management Company",
            "Preferred Payment Method", "Payment Terms",
        ],
        "example": [
            "ERN001", "Main Office", 101, "ABC Properties",
            "Jane Doe", "Property Manager", "jane@abc.com", "555-0200",
            "456 Oak Ave, Springfield IL", "portal.abc.com", "V-1234",
            "Primary landlord", "456 Oak Ave", "456 Oak Ave", "Suite 100",
            "Springfield", "IL", "62701", "PO Box 900", "", "Springfield",
            "IL", "62702", "555-0201", "555-0202", "abcproperties.com",
            "LLC", "12-3456789", "ABC Management Co", "ACH", "Net 30",
        ],
    },
    "vendors": {
        "filename": "vendors_template.xlsx",
        "title": "Vendors",
        "headers": [
            "Company Name", "Services", "Contact Name", "Email",
            "Phone", "Address", "Preferred", "Office Numbers", "Notes",
            "Address Line 1", "Address Line 2", "City", "State", "Zip Code",
            "1099 Vendor", "Tax ID", "Tax ID Type", "Legal Name",
            "Tax Classification", "Default Tax Box", "Default GL Account Code",
        ],
        "example": [
            "Acme Services", "HVAC Maintenance", "Bob Builder", "bob@acme.com",
            "555-0300", "789 Elm St", "Yes", "101;203;305",
            "Preferred HVAC vendor", "789 Elm St", "Unit 5", "Springfield",
            "IL", "62701", "Yes", "98-7654321", "ein", "Acme Services LLC",
            "llc", "nec_1", "6000",
        ],
    },
    "transitions": {
        "filename": "transitions_template.xlsx",
        "title": "Transitions",
        "headers": [
            "Office Number", "Transition Type", "Address", "New Address",
            "Status", "Sheet Name", "Lease Expiration", "Estimated Date", "Notes",
        ],
        "example": [
            101, "relocation", "123 Main St", "456 New Blvd",
            "in_progress", "Sheet1", "12/31/2026", "06/01/2026",
            "Moving to larger space",
        ],
    },
    "hvac-contracts": {
        "filename": "hvac_contracts_template.xlsx",
        "title": "HVAC Contracts",
        "headers": [
            "Office Number", "Office Name", "HVAC Company", "Contact",
            "Comments", "Frequency", "Last Serviced", "Next Service",
            "Manager", "Landlord Handles",
        ],
        "example": [
            101, "Main Office", "Cool Air Inc", "Mike Tech",
            "Annual contract", "Quarterly", "01/15/2026", "04/15/2026",
            "John Smith", "No",
        ],
    },
}

VALID_ENTITIES = set(TEMPLATES.keys())


def _generate_template(entity: str) -> io.BytesIO:
    tmpl = TEMPLATES[entity]
    wb = Workbook()
    ws = wb.active
    ws.title = tmpl["title"][:31]

    bold = Font(bold=True)
    example_font = Font(italic=True, color="999999")
    example_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    # Row 1: headers
    for col_idx, header in enumerate(tmpl["headers"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    # Row 2: example data
    for col_idx, val in enumerate(tmpl["example"], start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = example_font
        cell.fill = example_fill

    # Auto-size columns
    for col_idx, header in enumerate(tmpl["headers"], start=1):
        max_len = len(str(header))
        if col_idx - 1 < len(tmpl["example"]) and tmpl["example"][col_idx - 1] is not None:
            max_len = max(max_len, len(str(tmpl["example"][col_idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/{entity}/template")
async def download_template(
    entity: str,
    _=Depends(require_role("admin", "editor")),
):
    if entity not in VALID_ENTITIES:
        raise HTTPException(status_code=404, detail=f"Unknown entity: {entity}")
    tmpl = TEMPLATES[entity]
    buf = _generate_template(entity)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{tmpl["filename"]}"'},
    )


@router.post("/{entity}/import")
async def import_data(
    entity: str,
    file: UploadFile = File(...),
    force: bool = Query(
        default=False,
        description="Re-apply a file that has already been imported.",
    ),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role("admin", "editor")),
):
    if entity not in VALID_ENTITIES:
        raise HTTPException(status_code=404, detail=f"Unknown entity: {entity}")

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="File must be an XLSX file")

    contents = await file.read()
    org_id = current_user.organization_id

    # Guard against the most common bulk-load accident: the same spreadsheet
    # being uploaded twice. Callers can still override deliberately.
    if force:
        fingerprint = content_fingerprint(entity, contents)
    else:
        try:
            fingerprint = await import_assurance.check_replay(
                db, organization_id=org_id, entity_type=entity, payload=contents
            )
        except import_assurance.ReplayDetected as e:
            raise HTTPException(
                status_code=409,
                detail={
                    "message": str(e),
                    "previous_batch_id": str(e.batch.id),
                    "hint": "Re-send with ?force=true to import it again.",
                },
            )

    importer = IMPORTERS[entity]
    try:
        result = await importer(db, contents)
    except Exception as e:
        await import_assurance.record_batch(
            db,
            organization_id=org_id,
            source="xlsx",
            entity_type=entity,
            fingerprint=fingerprint,
            file_name=file.filename,
            created=0,
            updated=0,
            skipped=0,
            errors=[str(e)],
            imported_by_id=current_user.id,
            status="failed",
        )
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")

    batch = await import_assurance.record_batch(
        db,
        organization_id=org_id,
        source="xlsx",
        entity_type=entity,
        fingerprint=fingerprint,
        file_name=file.filename,
        created=result.created,
        updated=result.updated,
        skipped=result.skipped,
        errors=result.errors,
        imported_by_id=current_user.id,
    )

    payload = result.to_dict()
    payload["batch_id"] = str(batch.id)
    return payload


@router.get("/batches")
async def list_import_batches(
    entity: str | None = Query(default=None),
    limit: int = Query(default=50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role("admin", "editor")),
):
    """History of bulk loads, including the rows that failed in each."""
    stmt = (
        select(ImportBatch)
        .where(ImportBatch.organization_id == current_user.organization_id)
        .order_by(ImportBatch.created_at.desc())
        .limit(limit)
    )
    if entity:
        stmt = stmt.where(ImportBatch.entity_type == entity)
    batches = (await db.execute(stmt)).scalars().all()
    return [
        {
            "id": str(b.id),
            "source": b.source,
            "entity_type": b.entity_type,
            "file_name": b.file_name,
            "status": b.status,
            "rows_total": b.rows_total,
            "created_count": b.created_count,
            "updated_count": b.updated_count,
            "skipped_count": b.skipped_count,
            "error_count": b.error_count,
            "row_errors": b.row_errors,
            "created_at": b.created_at.isoformat(),
        }
        for b in batches
    ]


@router.post("/tie-out")
async def import_tie_out(
    source_counts: dict[str, int],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role("admin", "editor")),
):
    """Compare source-system record counts against what actually landed."""
    return await import_assurance.build_tie_out(
        db,
        organization_id=current_user.organization_id,
        source_counts=source_counts,
    )
