import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


def generate_xlsx(title: str, headers: list[str], rows: list[list[Any]]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet names max 31 chars

    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value="" if value is None else value)

    # Auto-size columns
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row) and row[col_idx - 1] is not None:
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
