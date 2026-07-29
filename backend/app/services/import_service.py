"""Import service: upsert logic for bulk XLSX imports of each entity."""

from __future__ import annotations

import re
import uuid
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select, delete, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.office import Office, Manager
from app.models.lease import Lease
from app.schemas.lease import normalize_lease_status
from app.models.landlord import Landlord
from app.models.vendor import Vendor, vendor_offices
from app.models.general_ledger import GLAccount
from app.models.transition import OfficeTransition
from app.models.hvac_contract import HvacContract


# ── helpers ──────────────────────────────────────────────────────────

_NON_DATE_STRINGS = frozenset({
    "??", "?", "n/a", "na", "tbd", "none", "---", "--", "-", "updated",
    "see lease", "per lease", "ongoing", "month to month", "mtm",
})


def safe_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def safe_int(val: Any) -> int | None:
    if val is None:
        return None
    try:
        return int(float(str(val).strip().replace("#", "")))
    except (ValueError, TypeError):
        return None


def safe_bool(val: Any) -> bool:
    if val is None:
        return False
    s = str(val).strip().lower()
    return s in ("yes", "true", "1", "y")


def safe_decimal(val: Any) -> Decimal | None:
    """Parse a numeric/currency value into a Decimal, tolerating $ and commas."""
    if val is None:
        return None
    if isinstance(val, Decimal):
        return val
    s = str(val).strip().replace("$", "").replace(",", "")
    if not s or s.lower() in _NON_DATE_STRINGS:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def safe_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        d = val.date()
        return d if 1900 <= d.year <= 2100 else None
    if isinstance(val, date):
        return val if 1900 <= val.year <= 2100 else None
    s = str(val).strip()
    if not s or s.lower() in _NON_DATE_STRINGS:
        return None
    date_candidate = s.split(",")[0].strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y",
                "%B %d, %Y", "%b %d, %Y", "%B %Y", "%b %Y"):
        try:
            d = datetime.strptime(date_candidate, fmt).date()
            return d if 1900 <= d.year <= 2100 else None
        except ValueError:
            continue
    try:
        from dateutil.parser import parse as _dp
        d = _dp(date_candidate, fuzzy=False).date()
        return d if 1900 <= d.year <= 2100 else None
    except Exception:
        pass
    return None


def parse_notice_days(notice_str: str | None) -> int | None:
    if not notice_str:
        return None
    match = re.search(r"(\d+)", str(notice_str))
    return int(match.group(1)) if match else None


def _is_example_row(values: list, example_markers: list[str]) -> bool:
    """Check if row looks like our template example row."""
    first = safe_str(values[0]) if values else None
    if first and first.lower() in [m.lower() for m in example_markers]:
        return True
    return False


def _is_row_empty(values: list) -> bool:
    return all(v is None or str(v).strip() == "" for v in values)


def _read_rows(file_bytes: bytes, example_first_cell: str | None = None) -> list[dict[str, Any]]:
    """Read XLSX bytes → list of dicts keyed by header names."""
    import io
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows()
    header_row = next(rows_iter, None)
    if header_row is None:
        return []
    headers = [safe_str(c.value) or f"col_{i}" for i, c in enumerate(header_row)]

    result: list[dict[str, Any]] = []
    for row in rows_iter:
        values = [c.value for c in row]
        if _is_row_empty(values):
            continue
        # Skip example row
        if example_first_cell and safe_str(values[0]) and \
                safe_str(values[0]).lower() == example_first_cell.lower():
            continue
        result.append({h: v for h, v in zip(headers, values)})
    return result


# ── lookup caches ────────────────────────────────────────────────────

async def _build_office_number_map(db: AsyncSession) -> dict[int, uuid.UUID]:
    """Map office_number → id for active offices."""
    stmt = select(Office.office_number, Office.id).where(Office.is_deleted.is_(False))
    rows = (await db.execute(stmt)).all()
    return {r[0]: r[1] for r in rows}


async def _build_manager_name_map(db: AsyncSession) -> dict[str, uuid.UUID]:
    """Map lowercase manager name → id."""
    stmt = select(Manager.name, Manager.id)
    rows = (await db.execute(stmt)).all()
    return {r[0].lower(): r[1] for r in rows}


async def _build_gl_account_code_map(db: AsyncSession) -> dict[str, uuid.UUID]:
    """Map GL account code → id for active chart-of-accounts entries."""
    stmt = select(GLAccount.code, GLAccount.id).where(GLAccount.is_active.is_(True))
    rows = (await db.execute(stmt)).all()
    return {str(r[0]).strip(): r[1] for r in rows}


# ── entity importers ────────────────────────────────────────────────

class ImportResult:
    def __init__(self):
        self.created = 0
        self.updated = 0
        self.skipped = 0
        self.errors: list[str] = []

    def to_dict(self):
        return {
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "errors": self.errors,
        }


async def import_managers(db: AsyncSession, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    rows = _read_rows(file_bytes, example_first_cell="John Smith")

    existing_map: dict[str, Manager] = {}
    stmt = select(Manager)
    for mgr in (await db.scalars(stmt)).all():
        existing_map[mgr.name.lower()] = mgr

    for i, row in enumerate(rows, start=2):
        name = safe_str(row.get("Name"))
        if not name:
            result.errors.append(f"Row {i}: Name is required")
            continue
        key = name.lower()
        if key in existing_map:
            mgr = existing_map[key]
            mgr.email = safe_str(row.get("Email")) or mgr.email
            mgr.phone = safe_str(row.get("Phone")) or mgr.phone
            result.updated += 1
        else:
            mgr = Manager(
                name=name,
                email=safe_str(row.get("Email")),
                phone=safe_str(row.get("Phone")),
            )
            db.add(mgr)
            existing_map[key] = mgr
            result.created += 1

    await db.commit()
    return result


async def import_offices(db: AsyncSession, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    rows = _read_rows(file_bytes, example_first_cell="101")
    manager_map = await _build_manager_name_map(db)
    gl_map = await _build_gl_account_code_map(db)

    existing_map: dict[int, Office] = {}
    stmt = select(Office).where(Office.is_deleted.is_(False))
    for off in (await db.scalars(stmt)).all():
        existing_map[off.office_number] = off

    for i, row in enumerate(rows, start=2):
        office_number = safe_int(row.get("Office Number"))
        location_type = safe_str(row.get("Location Type"))
        location_name = safe_str(row.get("Location Name"))
        if not office_number or not location_type or not location_name:
            result.errors.append(f"Row {i}: Office Number, Location Type, and Location Name are required")
            continue

        manager_id = None
        mgr_name = safe_str(row.get("Manager"))
        if mgr_name:
            manager_id = manager_map.get(mgr_name.lower())

        gl_account_id = None
        gl_code = safe_str(row.get("GL Account Code"))
        if gl_code:
            gl_account_id = gl_map.get(gl_code.strip())

        fields = dict(
            region_number=safe_int(row.get("Region Number")),
            location_type=location_type,
            location_name=location_name,
            manager_id=manager_id,
            is_active=not safe_bool(row.get("Active")) if safe_str(row.get("Active")) and safe_str(row.get("Active")).lower() == "no" else True if safe_str(row.get("Active")) is None else safe_bool(row.get("Active")),
            address_line_1=safe_str(row.get("Address Line 1")),
            address_line_2=safe_str(row.get("Address Line 2")),
            city=safe_str(row.get("City")),
            state=safe_str(row.get("State")),
            zip_code=safe_str(row.get("Zip Code")),
            phone_number=safe_str(row.get("Phone")),
            fax=safe_str(row.get("Fax")),
            email=safe_str(row.get("Email")),
            mail_shipping=safe_str(row.get("Mail/Shipping")),
            sector=safe_str(row.get("Sector")),
            notes=safe_str(row.get("Notes")),
            other_names=safe_str(row.get("Other Names")),
            crown_property_on_site=safe_str(row.get("Crown Property On Site")),
            additional_info=safe_str(row.get("Additional Info")),
            closing_notes=safe_str(row.get("Closing Notes")),
            total_sqft=safe_decimal(row.get("Total SqFt")),
            usable_sqft=safe_decimal(row.get("Usable SqFt")),
            headcount_capacity=safe_int(row.get("Headcount Capacity")),
            current_headcount=safe_int(row.get("Current Headcount")),
            space_type=safe_str(row.get("Space Type")),
            owner_same_as_landlord=safe_bool(row.get("Owner Same As Landlord")),
            owner_name=safe_str(row.get("Owner Name")),
            owner_company=safe_str(row.get("Owner Company")),
            owner_email=safe_str(row.get("Owner Email")),
            owner_phone=safe_str(row.get("Owner Phone")),
            owner_address_line_1=safe_str(row.get("Owner Address Line 1")),
            owner_address_line_2=safe_str(row.get("Owner Address Line 2")),
            owner_city=safe_str(row.get("Owner City")),
            owner_state=safe_str(row.get("Owner State")),
            owner_zip_code=safe_str(row.get("Owner Zip Code")),
            gl_account_id=gl_account_id,
        )

        if office_number in existing_map:
            off = existing_map[office_number]
            for k, v in fields.items():
                setattr(off, k, v)
            result.updated += 1
        else:
            off = Office(office_number=office_number, **fields)
            db.add(off)
            existing_map[office_number] = off
            result.created += 1

    await db.commit()
    return result


async def import_leases(db: AsyncSession, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    rows = _read_rows(file_bytes, example_first_cell="101 - Main Office")
    office_map = await _build_office_number_map(db)
    manager_map = await _build_manager_name_map(db)

    existing_map: dict[str, Lease] = {}
    stmt = select(Lease).where(Lease.is_deleted.is_(False))
    for lease in (await db.scalars(stmt)).all():
        existing_map[lease.lease_name.lower()] = lease

    for i, row in enumerate(rows, start=2):
        lease_name = safe_str(row.get("Lease Name"))
        exp_year = safe_int(row.get("Expiration Year"))
        if not lease_name or exp_year is None:
            result.errors.append(f"Row {i}: Lease Name and Expiration Year are required")
            continue

        office_id = None
        off_num = safe_int(row.get("Office Number"))
        if off_num:
            office_id = office_map.get(off_num)

        manager_id = None
        mgr_name = safe_str(row.get("Manager"))
        if mgr_name:
            manager_id = manager_map.get(mgr_name.lower())

        fields = dict(
            office_id=office_id,
            manager_id=manager_id,
            lease_expiration=safe_date(row.get("Expiration Date")),
            lessor_name=safe_str(row.get("Lessor Name")),
            notice_period=safe_str(row.get("Notice Period")),
            notice_period_days=parse_notice_days(safe_str(row.get("Notice Days"))) if safe_str(row.get("Notice Days")) else safe_int(row.get("Notice Days")),
            lease_notice_date=safe_date(row.get("Notice Date")),
            notice_given_date=safe_date(row.get("Notice Given Date")),
            status=normalize_lease_status(row.get("Status")),
            expiration_year=exp_year,
            lease_commencement_date=safe_date(row.get("Lease Commencement Date")),
            accounting_standard=safe_str(row.get("Accounting Standard")),
            lease_classification=safe_str(row.get("Lease Classification")),
            payment_amount=safe_decimal(row.get("Payment Amount")),
            payment_frequency=safe_str(row.get("Payment Frequency")),
            annual_escalation_rate=safe_decimal(row.get("Annual Escalation Rate")),
            incremental_borrowing_rate=safe_decimal(row.get("Incremental Borrowing Rate")),
            initial_direct_costs=safe_decimal(row.get("Initial Direct Costs")),
            lease_incentives=safe_decimal(row.get("Lease Incentives")),
            prepaid_rent=safe_decimal(row.get("Prepaid Rent")),
            residual_value_guarantee=safe_decimal(row.get("Residual Value Guarantee")),
            is_short_term_lease=safe_bool(row.get("Short Term Lease")),
            is_low_value_lease=safe_bool(row.get("Low Value Lease")),
            currency=safe_str(row.get("Currency")) or "USD",
        )

        key = lease_name.lower()
        if key in existing_map:
            lease = existing_map[key]
            for k, v in fields.items():
                setattr(lease, k, v)
            result.updated += 1
        else:
            lease = Lease(lease_name=lease_name, **fields)
            db.add(lease)
            existing_map[key] = lease
            result.created += 1

    await db.commit()
    return result


async def import_landlords(db: AsyncSession, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    rows = _read_rows(file_bytes, example_first_cell="ERN001")
    office_map = await _build_office_number_map(db)

    existing_map: dict[str, Landlord] = {}
    stmt = select(Landlord).where(Landlord.is_deleted.is_(False))
    for ll in (await db.scalars(stmt)).all():
        if ll.ern:
            existing_map[f"ern:{ll.ern.lower()}"] = ll
        elif ll.landlord_company and ll.contact_name:
            existing_map[f"comp:{ll.landlord_company.lower()}|{ll.contact_name.lower()}"] = ll

    for i, row in enumerate(rows, start=2):
        ern = safe_str(row.get("ERN"))
        company = safe_str(row.get("Landlord Company"))
        contact = safe_str(row.get("Contact Name"))

        if not ern and not company:
            result.errors.append(f"Row {i}: ERN or Landlord Company is required")
            continue

        office_id = None
        off_num = safe_int(row.get("Office Number"))
        if off_num:
            office_id = office_map.get(off_num)

        fields = dict(
            ern=ern,
            office_name=safe_str(row.get("Office Name")),
            office_id=office_id,
            landlord_company=company,
            contact_name=contact,
            title=safe_str(row.get("Title")),
            contact_email=safe_str(row.get("Email")),
            contact_phone=safe_str(row.get("Phone")),
            contact_mailing_address=safe_str(row.get("Mailing Address")),
            online_sign_in=safe_str(row.get("Online Sign In")),
            vendor_id=safe_str(row.get("Vendor ID")),
            notes=safe_str(row.get("Notes")),
            address=safe_str(row.get("Address")),
            address_line_1=safe_str(row.get("Address Line 1")),
            address_line_2=safe_str(row.get("Address Line 2")),
            city=safe_str(row.get("City")),
            state=safe_str(row.get("State")),
            zip_code=safe_str(row.get("Zip Code")),
            mailing_address_line_1=safe_str(row.get("Mailing Address Line 1")),
            mailing_address_line_2=safe_str(row.get("Mailing Address Line 2")),
            mailing_city=safe_str(row.get("Mailing City")),
            mailing_state=safe_str(row.get("Mailing State")),
            mailing_zip_code=safe_str(row.get("Mailing Zip Code")),
            secondary_phone=safe_str(row.get("Secondary Phone")),
            fax=safe_str(row.get("Fax")),
            website=safe_str(row.get("Website")),
            entity_type=safe_str(row.get("Entity Type")),
            tax_id=safe_str(row.get("Tax ID")),
            management_company=safe_str(row.get("Management Company")),
            preferred_payment_method=safe_str(row.get("Preferred Payment Method")),
            payment_terms=safe_str(row.get("Payment Terms")),
        )

        # Match key
        key = None
        if ern:
            key = f"ern:{ern.lower()}"
        elif company and contact:
            key = f"comp:{company.lower()}|{contact.lower()}"

        if key and key in existing_map:
            ll = existing_map[key]
            for k, v in fields.items():
                setattr(ll, k, v)
            result.updated += 1
        else:
            ll = Landlord(**fields)
            db.add(ll)
            if key:
                existing_map[key] = ll
            result.created += 1

    await db.commit()
    return result


async def import_vendors(db: AsyncSession, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    rows = _read_rows(file_bytes, example_first_cell="Acme Services")
    office_map = await _build_office_number_map(db)
    gl_map = await _build_gl_account_code_map(db)

    existing_map: dict[str, Vendor] = {}
    stmt = select(Vendor).where(Vendor.is_deleted.is_(False))
    for v in (await db.scalars(stmt)).all():
        existing_map[v.company_name.lower()] = v

    for i, row in enumerate(rows, start=2):
        company_name = safe_str(row.get("Company Name"))
        if not company_name:
            result.errors.append(f"Row {i}: Company Name is required")
            continue

        default_gl_account_id = None
        gl_code = safe_str(row.get("Default GL Account Code"))
        if gl_code:
            default_gl_account_id = gl_map.get(gl_code.strip())

        fields = dict(
            services=safe_str(row.get("Services")),
            contact_name=safe_str(row.get("Contact Name")),
            contact_email=safe_str(row.get("Email")),
            contact_phone=safe_str(row.get("Phone")),
            address=safe_str(row.get("Address")),
            is_preferred=safe_bool(row.get("Preferred")),
            notes=safe_str(row.get("Notes")),
            address_line_1=safe_str(row.get("Address Line 1")),
            address_line_2=safe_str(row.get("Address Line 2")),
            city=safe_str(row.get("City")),
            state=safe_str(row.get("State")),
            zip_code=safe_str(row.get("Zip Code")),
            is_1099_vendor=safe_bool(row.get("1099 Vendor")),
            tax_id=safe_str(row.get("Tax ID")),
            tax_id_type=safe_str(row.get("Tax ID Type")),
            legal_name=safe_str(row.get("Legal Name")),
            tax_classification=safe_str(row.get("Tax Classification")),
            default_tax_box=safe_str(row.get("Default Tax Box")),
            default_gl_account_id=default_gl_account_id,
        )

        key = company_name.lower()
        if key in existing_map:
            vendor = existing_map[key]
            for k, v in fields.items():
                setattr(vendor, k, v)
            result.updated += 1
        else:
            vendor = Vendor(company_name=company_name, **fields)
            db.add(vendor)
            existing_map[key] = vendor
            result.created += 1

        # Sync offices
        await db.flush()  # ensure vendor.id exists
        office_nums_str = safe_str(row.get("Office Numbers"))
        if office_nums_str:
            nums = [safe_int(n.strip()) for n in office_nums_str.split(";") if n.strip()]
            oids = [office_map[n] for n in nums if n and n in office_map]
            await db.execute(delete(vendor_offices).where(vendor_offices.c.vendor_id == vendor.id))
            for oid in oids:
                await db.execute(vendor_offices.insert().values(vendor_id=vendor.id, office_id=oid))

    await db.commit()
    return result


async def import_transitions(db: AsyncSession, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    rows = _read_rows(file_bytes, example_first_cell="101")
    office_map = await _build_office_number_map(db)

    existing_map: dict[str, OfficeTransition] = {}
    stmt = select(OfficeTransition).where(OfficeTransition.is_deleted.is_(False))
    for t in (await db.scalars(stmt)).all():
        if t.office_number and t.transition_type:
            existing_map[f"{t.office_number}|{t.transition_type.lower()}"] = t

    for i, row in enumerate(rows, start=2):
        off_num = safe_int(row.get("Office Number"))
        ttype = safe_str(row.get("Transition Type"))
        if not ttype:
            result.errors.append(f"Row {i}: Transition Type is required")
            continue

        office_id = None
        if off_num:
            office_id = office_map.get(off_num)

        fields = dict(
            office_id=office_id,
            office_number=off_num,
            transition_type=ttype,
            address=safe_str(row.get("Address")),
            new_address=safe_str(row.get("New Address")),
            status=safe_str(row.get("Status")) or "in_progress",
            sheet_name=safe_str(row.get("Sheet Name")),
            lease_expiration=safe_str(row.get("Lease Expiration")),
            estimated_date=safe_str(row.get("Estimated Date")),
            notes=safe_str(row.get("Notes")),
        )

        key = f"{off_num}|{ttype.lower()}" if off_num else None
        if key and key in existing_map:
            t = existing_map[key]
            for k, v in fields.items():
                setattr(t, k, v)
            result.updated += 1
        else:
            t = OfficeTransition(**fields)
            db.add(t)
            if key:
                existing_map[key] = t
            result.created += 1

    await db.commit()
    return result


async def import_hvac_contracts(db: AsyncSession, file_bytes: bytes) -> ImportResult:
    result = ImportResult()
    rows = _read_rows(file_bytes, example_first_cell="101")
    office_map = await _build_office_number_map(db)
    manager_map = await _build_manager_name_map(db)

    existing_map: dict[str, HvacContract] = {}
    stmt = select(HvacContract).where(HvacContract.is_deleted.is_(False))
    for hc in (await db.scalars(stmt)).all():
        if hc.office_number and hc.hvac_company:
            existing_map[f"{hc.office_number}|{hc.hvac_company.lower()}"] = hc

    for i, row in enumerate(rows, start=2):
        off_num = safe_int(row.get("Office Number"))
        hvac_co = safe_str(row.get("HVAC Company"))

        office_id = None
        if off_num:
            office_id = office_map.get(off_num)

        manager_id = None
        mgr_name = safe_str(row.get("Manager"))
        if mgr_name:
            manager_id = manager_map.get(mgr_name.lower())

        last_serviced = safe_str(row.get("Last Serviced"))
        next_service = safe_str(row.get("Next Service"))

        fields = dict(
            office_id=office_id,
            office_number=off_num,
            office_name=safe_str(row.get("Office Name")),
            hvac_company=hvac_co,
            contact=safe_str(row.get("Contact")),
            comments=safe_str(row.get("Comments")),
            frequency=safe_str(row.get("Frequency")),
            last_serviced=last_serviced,
            last_serviced_date=safe_date(last_serviced),
            next_service=next_service,
            next_service_date=safe_date(next_service),
            manager_id=manager_id,
            landlord_handles=safe_bool(row.get("Landlord Handles")),
        )

        key = f"{off_num}|{hvac_co.lower()}" if off_num and hvac_co else None
        if key and key in existing_map:
            hc = existing_map[key]
            for k, v in fields.items():
                setattr(hc, k, v)
            result.updated += 1
        else:
            hc = HvacContract(**fields)
            db.add(hc)
            if key:
                existing_map[key] = hc
            result.created += 1

    await db.commit()
    return result


# ── dispatcher ───────────────────────────────────────────────────────

IMPORTERS = {
    "managers": import_managers,
    "offices": import_offices,
    "leases": import_leases,
    "landlords": import_landlords,
    "vendors": import_vendors,
    "transitions": import_transitions,
    "hvac-contracts": import_hvac_contracts,
}
